import csv
import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import tkcalendar as tkc
from db import alterar_produto_db, cadastrar_produto, excluir_produto_db, is_numeric, obter_produto, obter_todos_produtos, obter_todos_produtos_cadastrados
from operations import backup_manual, cadastrar_venda
from ttkthemes import ThemedStyle, ThemedTk
from PIL import Image, ImageTk



from operations import cadastrar_venda, exportar_vendas_para_csv, obter_vendas, excluir_venda_db, alterar_venda_db, gerar_relatorio_dia

def atualizar_data_hora(label):
    data_hora_atual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    label.config(text=f"Hoje é: {data_hora_atual}")
    label.after(1000, atualizar_data_hora, label)

def exibir_relatorio_dia(vendas_dia):
    relatorio_window = tk.Toplevel()
    relatorio_window.title("Relatório do Dia")

    tree = ttk.Treeview(relatorio_window, columns=("ID", "Item", "Preço (R$)", "Quantidade", "Tipo de Quantidade", "Tipo de Pagamento", "Data da Venda"))
    # Adicione cabeçalhos e dados à tabela
    # Código para preencher a tabela com as vendas do dia
    tree.pack(expand=True, fill="both")

    relatorio_window.mainloop()

def cadastrar_venda_ui():
    def cadastrar_venda_handler():
        item = item_entry.get()
        preco = preco_entry.get().replace(',', '.')  # Substituir vírgula por ponto
        quantidade = quantidade_entry.get()
        tipo_quantidade = tipo_quantidade_combobox.get()
        tipo_pagamento = tipo_pagamento_combobox.get()

        if item and preco and quantidade and tipo_quantidade and tipo_pagamento:
            cadastrar_venda(item, float(preco), quantidade, tipo_quantidade, tipo_pagamento)
            messagebox.showinfo("Sucesso", "Venda cadastrada com sucesso!")
            cadastrar_venda_window.destroy()
        else:
            messagebox.showerror("Erro", "Por favor, preencha todos os campos.")

    # Obter a lista de produtos cadastrados
    produtos_cadastrados = obter_todos_produtos_cadastrados()  # Esta função deve retornar uma lista de nomes de produtos

    cadastrar_venda_window = ThemedTk(theme="breeze")  # Usar ThemedTk com o tema atual
    cadastrar_venda_window.title("Cadastrar Venda")

    # Definir a geometria para centralizar na tela
    largura_janela = 500
    altura_janela = 400
    largura_tela = cadastrar_venda_window.winfo_screenwidth()
    altura_tela = cadastrar_venda_window.winfo_screenheight()
    pos_x = (largura_tela - largura_janela) // 2
    pos_y = (altura_tela - altura_janela) // 2
    cadastrar_venda_window.geometry(f'{largura_janela}x{altura_janela}+{pos_x}+{pos_y}')

    # Restante do código para criar os widgets e definir o layout da janela

    fonte = ("Roboto", 12)

    tk.Label(cadastrar_venda_window, text="Produto:", font=fonte).pack()
    item_entry = ttk.Combobox(cadastrar_venda_window, values=produtos_cadastrados, font=fonte)
    item_entry.pack()

    tk.Label(cadastrar_venda_window, text="Preço (R$):", font=fonte).pack()
    preco_entry = tk.Entry(cadastrar_venda_window, font=fonte)
    preco_entry.pack()

    tk.Label(cadastrar_venda_window, text="Quantidade:", font=fonte).pack()
    quantidade_entry = tk.Entry(cadastrar_venda_window, font=fonte)
    quantidade_entry.pack()

    tk.Label(cadastrar_venda_window, text="Tipo de Quantidade:", font=fonte).pack()
    tipo_quantidade_combobox = ttk.Combobox(cadastrar_venda_window, values=["Kg", "Gramas", "Litros", "Metros"], font=fonte)
    tipo_quantidade_combobox.pack()

    tk.Label(cadastrar_venda_window, text="Tipo de Pagamento:", font=fonte).pack()
    tipo_pagamento_combobox = ttk.Combobox(cadastrar_venda_window, values=["Pix", "Débito", "Crédito", "Dinheiro"], font=fonte)
    tipo_pagamento_combobox.pack()

    # Criar o botão Cadastrar
    botao_cadastrar = ttk.Button(cadastrar_venda_window, text="Cadastrar Venda", command=cadastrar_venda_handler, style="TButton")
    botao_cadastrar.place(relx=0.5, rely=0.9, anchor="center")

def cadastrar_produto_ui():
    def cadastrar_produto_handler():
        nome = nome_entry.get().capitalize()
        quantidade = quantidade_entry.get()
        preco = preco_entry.get()

        if nome and quantidade and preco:
            produto = obter_produto(nome, quantidade, preco)
            if produto:
                messagebox.showerror("Erro", "Produto já cadastrado!")
            else:
                cadastrar_produto(nome, quantidade, preco)
                messagebox.showinfo("Sucesso", "Produto cadastrado com sucesso!")
                cadastrar_produto_window.destroy()
        else:
            messagebox.showerror("Erro", "Por favor, preencha todos os campos.")

    cadastrar_produto_window = ThemedTk(theme="breeze")  # Usar ThemedTk com o tema atual
    cadastrar_produto_window.title("Cadastrar Produto")

    # Definir a geometria para centralizar na tela
    largura_janela = 300
    altura_janela = 250
    largura_tela = cadastrar_produto_window.winfo_screenwidth()
    altura_tela = cadastrar_produto_window.winfo_screenheight()
    pos_x = (largura_tela - largura_janela) // 2
    pos_y = (altura_tela - altura_janela) // 2
    cadastrar_produto_window.geometry(f'{largura_janela}x{altura_janela}+{pos_x}+{pos_y}')

    fonte = ("Roboto", 12)

    tk.Label(cadastrar_produto_window, text="Nome do Produto:", font=fonte).pack()
    nome_entry = tk.Entry(cadastrar_produto_window, font=fonte)
    nome_entry.pack()

    tk.Label(cadastrar_produto_window, text="Quantidade:", font=fonte).pack()
    quantidade_entry = tk.Entry(cadastrar_produto_window, font=fonte)
    quantidade_entry.pack()

    tk.Label(cadastrar_produto_window, text="Preço:", font=fonte).pack()
    preco_entry = tk.Entry(cadastrar_produto_window, font=fonte)
    preco_entry.pack()

    # Criar o botão Cadastrar
    botao_cadastrar = ttk.Button(cadastrar_produto_window, text="Cadastrar Produto", command=cadastrar_produto_handler, style="TButton")
    botao_cadastrar.place(relx=0.5, rely=0.8, anchor="center")

    cadastrar_produto_window.mainloop()

def visualizar_estoque():
    produtos = obter_todos_produtos()

    def filtrar_produtos(event):
        filtro = filtro_entry.get().lower()
        produtos_filtrados = [produto for produto in produtos if filtro in produto[1].lower()]
        for row in tree.get_children():
            tree.delete(row)
        for produto in produtos_filtrados:
            preco_formatado = f"R${float(produto[3]):.2f}" if is_numeric(produto[3]) else "Preço não disponível"
            tree.insert("", "end", values=(produto[1], produto[2], preco_formatado))

    def excluir_produto():
        selected_item = tree.selection()
        if selected_item:
            produto_nome = tree.item(selected_item[0], "values")[0]
            excluir_produto_db(produto_nome)
            tree.delete(selected_item[0])
            messagebox.showinfo("Sucesso", "Produto excluído com sucesso!")

    def atualizar_tabela():
        for row in tree.get_children():
            tree.delete(row)
        produtos_atualizados = obter_todos_produtos()
        for produto in produtos_atualizados:
            preco_formatado = f"R${float(produto[3]):.2f}" if is_numeric(produto[3]) else "Preço não disponível"
            tree.insert("", "end", values=(produto[1], produto[2], preco_formatado))

    def alterar_produto(event=None):
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showerror("Erro", "Nenhum produto selecionado para alterar.")
            return

        produto_nome = tree.item(selected_item[0], "values")[0]
        produto = [p for p in produtos if p[1].lower() == produto_nome.lower()][0]

        def salvar_alteracoes():
            nome = nome_entry.get().capitalize()
            quantidade = quantidade_entry.get()
            preco = preco_entry.get().replace(',', '.')

            if nome and quantidade and preco:
                alterar_produto_db(produto[0], nome, quantidade, float(preco))
                messagebox.showinfo("Sucesso", "Produto alterado com sucesso!")
                alterar_produto_window.destroy()
                atualizar_tabela()
                estoque_window.lift()
            else:
                messagebox.showerror("Erro", "Por favor, preencha todos os campos.")

        alterar_produto_window = tk.Toplevel()
        alterar_produto_window.title("Alterar Produto")
        alterar_produto_window.geometry("400x300")

        fonte = ("Roboto", 12)

        tk.Label(alterar_produto_window, text="Nome do Produto:", font=fonte).pack()
        nome_entry = tk.Entry(alterar_produto_window, font=fonte)
        nome_entry.insert(0, produto[1])
        nome_entry.pack()

        tk.Label(alterar_produto_window, text="Quantidade:", font=fonte).pack()
        quantidade_entry = tk.Entry(alterar_produto_window, font=fonte)
        quantidade_entry.insert(0, produto[2])
        quantidade_entry.pack()

        tk.Label(alterar_produto_window, text="Preço (R$):", font=fonte).pack()
        preco_entry = tk.Entry(alterar_produto_window, font=fonte)
        preco_entry.insert(0, f"{produto[3]:.2f}".replace(".", ","))
        preco_entry.pack()

        ttk.Button(alterar_produto_window, text="Salvar Alterações", command=salvar_alteracoes, style='Custom.TButton').pack(side=tk.BOTTOM, pady=10, padx=5)

    estoque_window = ThemedTk(theme="breeze")
    estoque_window.title("Estoque de Produtos")

    # Configurar o estilo temático para os widgets ttk
    style = ThemedStyle(estoque_window)
    style.set_theme("breeze")

    largura_janela = 600
    altura_janela = 400
    largura_tela = estoque_window.winfo_screenwidth()
    altura_tela = estoque_window.winfo_screenheight()
    pos_x = (largura_tela - largura_janela) // 2
    pos_y = (altura_tela - altura_janela) // 2
    estoque_window.geometry(f'{largura_janela}x{altura_janela}+{pos_x}+{pos_y}')

    fonte = ("Roboto", 12)

    tree = ttk.Treeview(estoque_window, columns=("Produto", "Quantidade", "Preço"), show="headings")
    tree.heading("Produto", text="Produto")
    tree.heading("Quantidade", text="Quantidade")
    tree.heading("Preço", text="Preço")
    tree.pack(expand=True, fill=tk.BOTH)

    for produto in produtos:
        preco_formatado = f"R${float(produto[3]):.2f}" if is_numeric(produto[3]) else "Preço não disponível"
        tree.insert("", tk.END, values=(produto[1], produto[2], preco_formatado))

    tree.bind("<Double-1>", alterar_produto)

    frame_botoes = tk.Frame(estoque_window)
    frame_botoes.pack(pady=10)

    frame_busca = tk.Frame(frame_botoes)
    frame_busca.pack(side=tk.LEFT)

    lbl_busca = tk.Label(frame_busca, text="Buscar:", font=fonte)
    lbl_busca.pack(side=tk.LEFT)

    filtro_entry = tk.Entry(frame_busca, font=fonte)
    filtro_entry.pack(side=tk.LEFT, padx=5)
    filtro_entry.bind("<KeyRelease>", filtrar_produtos)

    # Criar estilo personalizado para os botões
    style.configure('Custom.TButton', font=fonte)

    btn_excluir = ttk.Button(frame_botoes, text="Excluir", command=excluir_produto, style='Custom.TButton')
    btn_excluir.pack(side=tk.LEFT, padx=5)

    btn_atualizar = ttk.Button(frame_botoes, text="Atualizar Tabela", command=atualizar_tabela, style='Custom.TButton')
    btn_atualizar.pack(side=tk.LEFT, padx=5)

    estoque_window.mainloop()

def centralizar_janela(janela):
    janela.update_idletasks()
    largura = janela.winfo_width()
    altura = janela.winfo_height()
    x = (janela.winfo_screenwidth() // 2) - (largura // 2)
    y = (janela.winfo_screenheight() // 2) - (altura // 2)
    janela.geometry(f'{largura}x{altura}+{x}+{y}')

def visualizar_vendas_ui():
    vendas = obter_vendas()

    def filtrar_vendas(event):
        filtro = filtro_entry.get().lower()
        vendas_filtradas = [venda for venda in vendas if filtro in venda[1].lower()]
        for row in tree.get_children():
            tree.delete(row)
        for venda in vendas_filtradas:
            preco_formatado = f"{venda[2]:.2f}".replace(".", ",")
            venda_formatada = (*venda[:2], preco_formatado, *venda[3:])
            tree.insert("", "end", values=venda_formatada)

    def excluir_venda():
        selected_item = tree.selection()
        if selected_item:
            venda_id = tree.item(selected_item[0], "values")[0]
            excluir_venda_db(venda_id)
            tree.delete(selected_item[0])
            messagebox.showinfo("Sucesso", "Venda excluída com sucesso!")

    def atualizar_tabela():
        for row in tree.get_children():
            tree.delete(row)
        vendas = obter_vendas()
        for venda in vendas:
            preco_formatado = f"{venda[2]:.2f}".replace(".", ",")
            venda_formatada = (*venda[:2], preco_formatado, *venda[3:])
            tree.insert("", "end", values=venda_formatada)

    def alterar_venda(event=None):
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showerror("Erro", "Nenhuma venda selecionada para alterar.")
            return

        venda_id = tree.item(selected_item[0], "values")[0]
        venda = [v for v in vendas if v[0] == int(venda_id)][0]

        def salvar_alteracoes():
            item = venda[1]
            preco = preco_entry.get().replace(',', '.')
            quantidade = quantidade_entry.get()
            tipo_quantidade = tipo_quantidade_combobox.get()
            tipo_pagamento = tipo_pagamento_combobox.get()

            if item and preco and quantidade and tipo_quantidade and tipo_pagamento:
                alterar_venda_db(venda_id, item, float(preco), quantidade, tipo_quantidade, tipo_pagamento)
                messagebox.showinfo("Sucesso", "Venda alterada com sucesso!")
                alterar_venda_window.destroy()
                atualizar_tabela()
                visualizar_vendas_window.lift()
                
            else:
                messagebox.showerror("Erro", "Por favor, preencha todos os campos.")
        
        alterar_venda_window = tk.Toplevel()
        alterar_venda_window.title("Alterar Venda")
        alterar_venda_window.geometry("400x300")

        style = ThemedStyle(alterar_venda_window)
        style.set_theme("breeze")

        fonte = ("Roboto", 12)

        # Definindo um estilo personalizado para labels
        style.configure('Custom.TLabel', font=fonte, foreground='#1E90FF')

        ttk.Label(alterar_venda_window, text="Item:", style='Custom.TLabel').pack()
        ttk.Label(alterar_venda_window, text=venda[1], font=fonte).pack()  # Exibir o nome do item como rótulo

        ttk.Label(alterar_venda_window, text="Preço (R$):", style='Custom.TLabel').pack()
        preco_entry = tk.Entry(alterar_venda_window, font=fonte)
        preco_entry.insert(0, f"{venda[2]:.2f}".replace(".", ","))
        preco_entry.pack()

        ttk.Label(alterar_venda_window, text="Quantidade:", style='Custom.TLabel').pack()
        quantidade_entry = tk.Entry(alterar_venda_window, font=fonte)
        quantidade_entry.insert(0, venda[3])
        quantidade_entry.pack()

        ttk.Label(alterar_venda_window, text="Tipo de Quantidade:", style='Custom.TLabel').pack()
        tipo_quantidade_combobox = ttk.Combobox(alterar_venda_window, values=["Kg", "Unidades", "Litros", "Metros"], font=fonte)
        tipo_quantidade_combobox.set(venda[4])
        tipo_quantidade_combobox.pack()

        ttk.Label(alterar_venda_window, text="Tipo de Pagamento:", style='Custom.TLabel').pack()
        tipo_pagamento_combobox = ttk.Combobox(alterar_venda_window, values=["Pix", "Débito", "Crédito", "Dinheiro"], font=fonte)
        tipo_pagamento_combobox.set(venda[5])
        tipo_pagamento_combobox.pack()

        style.configure('Custom.TButton', font=fonte)

        ttk.Button(alterar_venda_window, text="Salvar Alterações", command=salvar_alteracoes, style='Custom.TButton').pack(side=tk.BOTTOM, pady=10, padx=5)

        alterar_venda_window.update_idletasks()
        centralizar_janela(alterar_venda_window)

    visualizar_vendas_window = ThemedTk(theme="breeze")
    visualizar_vendas_window.title("Visualizar Vendas")
    visualizar_vendas_window.geometry("1000x400")

    style = ThemedStyle(visualizar_vendas_window)
    style.set_theme("breeze")

    fonte = ("Roboto", 12)

    tree = ttk.Treeview(visualizar_vendas_window, columns=("ID", "Produto", "Preço (R$)", "Quantidade", "Tipo de Quantidade", "Tipo de Pagamento", "Data da Venda"), show="headings")
    tree.heading("ID", text="ID")
    tree.heading("Produto", text="Produto")
    tree.heading("Preço (R$)", text="Preço (R$)")
    tree.heading("Quantidade", text="Quantidade")
    tree.heading("Tipo de Quantidade", text="Tipo de Quantidade")
    tree.heading("Tipo de Pagamento", text="Tipo de Pagamento")
    tree.heading("Data da Venda", text="Data da Venda")

    tree.column("ID", width=20)
    tree.column("Produto", width=150)
    tree.column("Preço (R$)", width=100)
    tree.column("Quantidade", width=100)
    tree.column("Tipo de Quantidade", width=120)
    tree.column("Tipo de Pagamento", width=120)
    tree.column("Data da Venda", width=150)

    for venda in vendas:
        preco_formatado = f"{venda[2]:.2f}".replace(".", ",")
        venda_formatada = (*venda[:2], preco_formatado, *venda[3:])
        tree.insert("", "end", values=venda_formatada)

    tree.bind("<Double-1>", alterar_venda)
    tree.pack(expand=True, fill="both")

    frame_botoes = tk.Frame(visualizar_vendas_window)
    frame_botoes.pack(pady=10)

    frame_busca = tk.Frame(frame_botoes)
    frame_busca.pack(side=tk.LEFT)

    lbl_busca = ttk.Label(frame_busca, text="Buscar:", style='Custom.TLabel')
    lbl_busca.pack(side=tk.LEFT)

    filtro_entry = tk.Entry(frame_busca, font=fonte)
    filtro_entry.pack(side=tk.LEFT, padx=5)
    filtro_entry.bind("<KeyRelease>", filtrar_vendas)

    style.configure('Custom.TButton', font=fonte)

    btn_excluir = ttk.Button(frame_botoes, text="Excluir", command=excluir_venda, style='Custom.TButton')
    btn_excluir.pack(side=tk.LEFT, padx=5)

    btn_atualizar = ttk.Button(frame_botoes, text="Atualizar Tabela", command=atualizar_tabela, style='Custom.TButton')
    btn_atualizar.pack(side=tk.LEFT, padx=5)

    visualizar_vendas_window.update_idletasks()
    centralizar_janela(visualizar_vendas_window)

    visualizar_vendas_window.mainloop()

def gerar_relatorio_dia():
    vendas = obter_vendas()
    today = datetime.today().strftime('%d-%m-%Y')
    today_display = datetime.today().strftime('%d/%m/%Y')
    vendas_dia = [venda for venda in vendas if venda[6].startswith(today_display)]

    # Criar uma nova janela para exibir o relatório do dia
    relatorio_window = tk.Toplevel()
    relatorio_window.title("Relatório do Dia")

    # Criar uma árvore para exibir as vendas do dia
    tree = ttk.Treeview(relatorio_window, columns=("ID", "Item", "Preço (R$)", "Quantidade", "Tipo de Quantidade", "Tipo de Pagamento", "Data da Venda"), show="headings")
    tree.heading("ID", text="ID")
    tree.heading("Item", text="Item")
    tree.heading("Preço (R$)", text="Preço (R$)")
    tree.heading("Quantidade", text="Quantidade")
    tree.heading("Tipo de Quantidade", text="Tipo de Quantidade")
    tree.heading("Tipo de Pagamento", text="Tipo de Pagamento")
    tree.heading("Data da Venda", text="Data da Venda")

    # Definir a largura das colunas
    tree.column("ID", width=50)
    tree.column("Item", width=150)
    tree.column("Preço (R$)", width=100)
    tree.column("Quantidade", width=100)
    tree.column("Tipo de Quantidade", width=150)
    tree.column("Tipo de Pagamento", width=150)
    tree.column("Data da Venda", width=150)


    # Adicionar as vendas do dia à árvore
    for venda in vendas_dia:
        preco_formatado = f"{venda[2]:.2f}".replace(".", ",")
        venda_formatada = (*venda[:2], preco_formatado, *venda[3:])
        tree.insert("", "end", values=venda_formatada)

    tree.pack(expand=True, fill="both")

    # Calcular o fechamento de caixa do dia
    total_vendido = sum(float(venda[2]) for venda in vendas_dia)  # Somar apenas os preços
    total_quantidade = sum(int(venda[3]) for venda in vendas_dia)  # Somar apenas as quantidades




    # Exibir o fechamento de caixa
    fechamento_texto = f"""
    Fechamento de Caixa do Dia:
    Total Vendido: R$ {total_vendido:.2f}

    """

    fechamento_label = tk.Label(relatorio_window, text=fechamento_texto, justify=tk.LEFT, font=("Roboto", 12))
    fechamento_label.pack(pady=10)

    def exportar_para_excel():
        # Diretório de Documentos do usuário
        diretorio_documentos = os.path.expanduser('~/Documents/relatorios')
        # Verificar se o diretório existe, se não, criar
        if not os.path.exists(diretorio_documentos):
            os.makedirs(diretorio_documentos)

        # Caminho completo para o arquivo Excel
        caminho_arquivo = os.path.join(diretorio_documentos, f'relatorio_dia_{today}.xlsx')

        # Criar um workbook e uma worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório de Vendas do Dia"

        # Adicionar cabeçalhos
        headers = ["ID", "Item", "Preço (R$)", "Quantidade", "Tipo de Quantidade", "Tipo de Pagamento", "Data da Venda"]
        ws.append(["Relatório de Vendas do Dia", today_display])
        ws.append([])  # Linha em branco
        ws.append(headers)

        # Adicionar as vendas do dia
        for venda in vendas_dia:
            ws.append(venda)

        # Adicionar resumo
        ws.append([])  # Linha em branco
        ws.append(["Resumo", "Valor"])
        ws.append(["Total Vendido no dia", f"R$ {total_vendido:.2f}"])

        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)  # Obter a letra da coluna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Salvar o arquivo
        wb.save(caminho_arquivo)

        messagebox.showinfo("Sucesso", f"Relatório exportado com sucesso!\nLocal: {caminho_arquivo}")

    # Adicionar botão para exportar para Excel
    exportar_button = ttk.Button(relatorio_window, text="Fechar Caixa e Exportar", command=exportar_para_excel)
    exportar_button.pack(pady=10)


def gerar_relatorio_por_dia(calendario):
    data_selecionada = calendario.get_date()
    vendas = obter_vendas()
    vendas_dia = [venda for venda in vendas if venda[6][:10] == data_selecionada]

    # Criar uma nova janela para exibir o relatório do dia
    relatorio_window = tk.Toplevel()
    relatorio_window.title(f"Relatório do Dia - {data_selecionada}")

    # Criar uma árvore para exibir as vendas do dia
    tree = ttk.Treeview(relatorio_window, columns=("ID", "Item", "Preço (R$)", "Quantidade", "Tipo de Quantidade", "Tipo de Pagamento", "Data da Venda"), show="headings")
    tree.heading("ID", text="ID")
    tree.heading("Item", text="Item")
    tree.heading("Preço (R$)", text="Preço (R$)")
    tree.heading("Quantidade", text="Quantidade")
    tree.heading("Tipo de Quantidade", text="Tipo de Quantidade")
    tree.heading("Tipo de Pagamento", text="Tipo de Pagamento")
    tree.heading("Data da Venda", text="Data da Venda")

    # Adicionar as vendas do dia à árvore
    for venda in vendas_dia:
        preco_formatado = f"{venda[2]:.2f}".replace(".", ",")
        venda_formatada = (*venda[:2], preco_formatado, *venda[3:])
        tree.insert("", "end", values=venda_formatada)

    # Ajustar a largura das colunas
    for col in tree['columns']:
        tree.column(col, width=150)  # Defina a largura desejada aqui    

    tree.pack(expand=True, fill="both")
# Abrir calendario clicando na data
def abrir_relatorio_por_dia(event):
    calendario = event.widget
    gerar_relatorio_por_dia(calendario)

    calendario.bind("<<CalendarSelected>>", abrir_relatorio_por_dia)
    
def exportar_vendas_ui():
    file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
    if file_path:
        exportar_vendas_para_csv(file_path)
        messagebox.showinfo("Sucesso", "Vendas exportadas com sucesso!")

def realizar_nova_venda(item, preco, quantidade, tipo_quantidade, tipo_pagamento):
    # Realizar a nova venda
    cadastrar_venda(item, preco, quantidade, tipo_quantidade, tipo_pagamento)

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


# Interface, Tela principal
def iniciar_interface():
    app = tk.Tk()
    app.title("Venda Smart")

    # Definir o ícone do aplicativo
    script_dir = os.path.dirname(os.path.abspath(__file__))
    icon_path = os.path.join(script_dir, 'assets', 'venda_smart_icon.ico')
    app.iconbitmap(icon_path)
    


    # Caminho absoluto para o diretório do script
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Caminho para a imagem no diretório 'assets'
    image_path = os.path.join(script_dir, 'assets', 'VendaSmart.png')

    # Tentar carregar e redimensionar a imagem
    try:
        logo_image = Image.open(image_path)
        logo_image = logo_image.resize((50, 50), Image.Resampling.LANCZOS)  # Redimensionar a imagem para 50x50 pixels
        logo_photo = ImageTk.PhotoImage(logo_image)
    except FileNotFoundError:
        print(f"Arquivo não encontrado: {image_path}")
        logo_photo = None
    except Exception as e:
        print(f"Erro ao abrir a imagem: {e}")
        logo_photo = None

    # Frame para o título e o logo
    frame_titulo = tk.Frame(app)
    frame_titulo.pack(pady=10)

    # Exibir a imagem do logo antes do título
    if logo_photo:
        label_logo = tk.Label(frame_titulo, image=logo_photo)
        label_logo.image = logo_photo  # Necessário para manter a referência da imagem
        label_logo.pack(side="left", padx=10)

    # Título
    titulo = tk.Label(frame_titulo, text="Venda Smart", font=("Roboto", 24))
    titulo.pack(side="left")
        
    # Criar um estilo temático
    style = ThemedStyle(app)
    style.set_theme("breeze")


    # Calcular a posição para centralizar a janela na tela
    largura_janela = 800
    altura_janela = 600
    largura_tela = app.winfo_screenwidth()
    altura_tela = app.winfo_screenheight()
    pos_x = (largura_tela - largura_janela) // 1
    pos_y = (altura_tela - altura_janela) // 1
    app.geometry(f'{largura_janela}x{altura_janela}+{pos_x}+{pos_y}')

    # Barra de Tarefas
    menubar = tk.Menu(app)
    app.config(menu=menubar)

    menu_acoes = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="Ações", menu=menu_acoes)
    menu_acoes.add_command(label="Cadastrar Venda", command=cadastrar_venda_ui)
    menu_acoes.add_command(label="Visualizar Vendas", command=visualizar_vendas_ui)

    menu_acoes = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="Produtos", menu=menu_acoes)
    menu_acoes.add_command(label="Cadastrar Produtos", command=cadastrar_produto_ui)
    menu_acoes.add_command(label="Visualizar Estoque", command=visualizar_estoque)

    menu_acoes_relatorios = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="Relatórios", menu=menu_acoes_relatorios)
    menu_acoes_relatorios.add_command(label="Relatório do dia", command=gerar_relatorio_dia)
    menu_acoes_relatorios.add_command(label="Exportar vendas", command=exportar_vendas_ui)
    menu_acoes_relatorios.add_command(label="Fazer Backup", command=backup_manual)

    menu_acoes = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="Opções", menu=menu_acoes)
    menu_acoes.add_command(label="Sair", command=app.quit)

 

    # Frame principal
    frame_principal = tk.Frame(app)
    frame_principal.pack(expand=True, fill='both')

    # Frames para colunas
    frame_coluna_esquerda = tk.Frame(frame_principal)
    frame_coluna_esquerda.grid(row=0, column=0, padx=10, pady=10)

    frame_coluna_direita = tk.Frame(frame_principal)
    frame_coluna_direita.grid(row=0, column=1, padx=10, pady=10)

    # Criar um frame para a parte inferior da interface
    frame_rodape = tk.Frame(app)
    frame_rodape.pack(side="bottom", fill="x")

    # Definir tamanho das colunas
    frame_principal.columnconfigure(0, weight=3)  # Primeira coluna
    frame_principal.columnconfigure(1, weight=1)  # Segunda coluna



    ##### COLUNA DA ESQUERDA ####

    # Botão Cadastrar Venda
    botao_cadastrar_venda = ttk.Button(frame_coluna_esquerda, text="Cadastrar Venda", command=cadastrar_venda_ui, width=30, padding=(10, 5))
    botao_cadastrar_venda.pack(pady=10)

    # Botão Visualizar Vendas
    botao_visualizar_vendas = ttk.Button(frame_coluna_esquerda, text="Visualizar Vendas", command=visualizar_vendas_ui, width=30, padding=(10, 5))
    botao_visualizar_vendas.pack(pady=10)

    # Adicionar um widget de calendário
    calendario = tkc.Calendar(frame_coluna_esquerda, locale="pt_BR", date_pattern="dd/mm/yyyy", selectmode="day")
    calendario.pack(pady=10)

    # Adicionar um botão para gerar o relatório do dia selecionado
    calendario.bind("<<CalendarSelected>>", abrir_relatorio_por_dia)

    # Botão gerar relatórios
    botao_relatorio_dia = ttk.Button(frame_coluna_esquerda, text="Visualizar Relatório do Dia", command=gerar_relatorio_dia, width=30, padding=(10, 5))
    botao_relatorio_dia.pack(pady=10)


    ##### COLUNA DA DIREITA ####
    

    # Botão Cadastro Estoque
    botao_backup = ttk.Button(frame_coluna_direita, text="Cadastrar Produto/Estoque", command=cadastrar_produto_ui, width=30, padding=(10, 5))
    botao_backup.pack(pady=10)

    #Botão Visualizar Estoque
    botao_visualizar_estoque = ttk.Button(frame_coluna_direita, text="Visualizar Estoque", command=visualizar_estoque,  width=30, padding=(10, 5))
    botao_visualizar_estoque.pack()
    
    # Botão Fazer Backup
    botao_backup = ttk.Button(frame_coluna_direita, text="Fazer Backup", command=backup_manual, width=30, padding=(10, 5))
    botao_backup.pack(pady=10)

    # Botão Sair
    botao_sair = ttk.Button(frame_coluna_direita, text="Sair", command=app.quit, width=30, padding=(10, 5))
    botao_sair.pack(pady=10)

    # Gráfico 

   # Rótulo para data e hora
    label_data_hora = tk.Label(frame_rodape, text="", font=("Roboto", 10))
    label_data_hora.pack(side=tk.BOTTOM,pady=5)
    atualizar_data_hora(label_data_hora)


    # Adicionar widgets ao frame do rodapé
    label_rodape = tk.Label(frame_rodape, text="T4W - © 2024 Todos os direitos reservados.", font=("Roboto", 10))
    label_rodape.pack(side=tk.BOTTOM, pady=5)
    atualizar_data_hora(label_data_hora)

    # Iniciar o loop principal da interface gráfica
    app.mainloop()




if __name__ == "__main__":
    iniciar_interface()


